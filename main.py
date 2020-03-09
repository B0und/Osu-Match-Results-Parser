import json
import os
import re
import sys
import webbrowser
from pprint import pprint

import requests
from lxml import html
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox

from des_ui import Ui_MainWindow


def get_username_from_id(user_id):
    url = "https://osu.ppy.sh/users/" + str(user_id)
    page = requests.get(url)
    tree = html.fromstring(page.content)
    results = tree.xpath("//script[@id='json-user']/text()")
    data = json.loads(str(results[0]))
    return data["username"]


def resize_cols(ws, index):
    # resize column to fit beatmap name but not too long
    new_len = len(ws.cell(row=1, column=index).value)
    if new_len > 100:
        ws.column_dimensions[get_column_letter(index)].width = 100
    else:
        ws.column_dimensions[get_column_letter(index)].width = new_len


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowTitle("Osu! Match Results Parser")
        app_icon = QtGui.QIcon()
        app_icon.addFile("icon.png", QtCore.QSize(256, 256))
        self.setWindowIcon(app_icon)

        self.ui.pushButton.clicked.connect(self.pushBtnFun)

    def pushBtnFun(self):
        url = self.ui.lineEditBeatmap.text()
        filename = self.ui.lineEditFilename.text()

        lobby_dict = {}

        # gather all scores from different urls in a dictionary

        url_list = url.split("|")
        for url in url_list:
            page = requests.get(url)
            tree = html.fromstring(page.content)
            results = tree.xpath("//script[@id='json-events']/text()")
            # print(dir(results[0]))
            # print(str(results[0]))
            data = json.loads(str(results[0]))

            for item in data["events"]:
                if "game" in item.keys():
                    for _ in item["game"].keys():
                        artist = item["game"]["beatmap"]["beatmapset"]["artist"]
                        song = item["game"]["beatmap"]["beatmapset"]["title"]
                        diff = item["game"]["beatmap"]["version"]
                        col_name_str = artist + " - " + song + f"[{diff}]"

                        score_list = []
                        for score in item["game"]["scores"]:
                            score_list.append(score)

                        if col_name_str in lobby_dict.keys():
                            current_list = lobby_dict[col_name_str]
                            lobby_dict[col_name_str] = current_list + score_list

                        else:
                            lobby_dict[col_name_str] = score_list

                        score_list = []
                        break

        lobby_dict_new = {}
        # sort all scores
        for k, v in lobby_dict.items():
            temp_list = []
            for item in v:
                username = get_username_from_id(item["user_id"])
                temp_list.append((item["score"], username))

            score_list = sorted(temp_list, key=lambda x: x[0], reverse=True)
            lobby_dict_new[k] = score_list

        # write scores into excel file
        wb = Workbook()
        ws = wb.active

        rowCnt = 1
        colCnt = 1

        for k, v in lobby_dict_new.items():
            ws.cell(row=rowCnt, column=colCnt, value=k)
            resize_cols(ws, colCnt)
            for item in v:
                rowCnt += 1
                ws.cell(row=rowCnt, column=colCnt, value=item[0])
                colCnt += 1
                ws.cell(row=rowCnt, column=colCnt, value=item[1])
                colCnt -= 1

            rowCnt = 1
            colCnt += 2

        wb.save(filename + ".xlsx")
        QMessageBox.about(self, "What did i just pick", "Done !")
        webbrowser.open(os.getcwd())
        self.close()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    root = MainWindow()
    root.show()
    sys.exit(app.exec_())
